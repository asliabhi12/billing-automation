import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from io import BytesIO
import json
from openpyxl.styles import Font

with open("customer_map.json", encoding="utf-8") as f:
    CUSTOMER_MAP = json.load(f)


def process_files(files):

    output = BytesIO()

    # ===== Extract Summary =====
    def extract_summary_data(file):
        file.seek(0)
        wb = load_workbook(file, data_only=True)
        ws = wb['Summary']

        def format_date(date_value):
            if isinstance(date_value, datetime):
                return date_value.strftime('%B %d, %Y')
            elif isinstance(date_value, str):
                try:
                    parsed_date = datetime.strptime(date_value, '%a, %b %d, %Y')
                    return parsed_date.strftime('%B %d, %Y')
                except ValueError:
                    return ''
            return ''

        start_date = format_date(ws['C8'].value)
        end_date = format_date(ws['C9'].value)

        subscription_id = ws['C5'].value.split('/')[-1]
        customer_name = CUSTOMER_MAP.get(subscription_id, "")

        return {
            "Customer Name": customer_name,
            "Start Date": start_date,
            "End Date": end_date,
            "Subscription ID": subscription_id
        }

    # ===== Process File =====
    def process_file(file_path, summary_data):
        file_path.seek(0)

        df = pd.read_excel(file_path, sheet_name='Data', engine="openpyxl")

        # Clean column names
        df.columns = df.columns.str.strip()

        # Rename columns
        df.rename(columns={
            'Meter': 'Meter Name',
            'ServiceName': 'Service Type',
            'ResourceLocation': 'Region',
            'ResourceType': 'Resource Name',
            'Cost': 'Total Cost'
        }, inplace=True)

        # Select required columns
        df_transformed = df[['Meter Name', 'Service Type', 'Resource Name', 'Region', 'Total Cost']].copy()

        # Clean & convert cost column
        df_transformed["Total Cost"] = (
            df_transformed["Total Cost"]
            .astype(str)
            .str.replace(",", "")
            .str.strip()
        )

        df_transformed["Total Cost"] = pd.to_numeric(df_transformed["Total Cost"], errors="coerce")
        df_transformed = df_transformed[df_transformed["Total Cost"].notna()]

        # Add summary data
        for col, val in summary_data.items():
            df_transformed[col] = val

        # Reorder columns
        df_transformed = df_transformed[[
            "Customer Name", "Start Date", "End Date", "Subscription ID",
            "Meter Name", "Service Type", "Resource Name", "Region", "Total Cost"
        ]]

        return df_transformed

    # ===== Write Excel =====
    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        from openpyxl.utils import get_column_letter

        for file in files:

            summary_data = extract_summary_data(file)
            df_transformed = process_file(file, summary_data)

            sheet_name = file.original_name.replace('.xlsx', '')
            sheet_name = sheet_name[:31].replace('/', '_').replace('\\', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_').replace('?', '_')

            df_transformed.to_excel(writer, sheet_name=sheet_name, index=False)

            worksheet = writer.sheets[sheet_name]

            # ===== Total Formula =====
            total_col_idx = df_transformed.columns.get_loc("Total Cost") + 1
            label_col_idx = total_col_idx - 1  # second last column

            total_col_letter = get_column_letter(total_col_idx)
            label_col_letter = get_column_letter(label_col_idx)

            data_last_row = len(df_transformed) + 1  # last data row in Excel

            # Place "Total Cost" in second last column
            worksheet[f"{label_col_letter}{data_last_row + 1}"] = "Total Cost"

            # Place SUM formula in last column
            worksheet[f"{total_col_letter}{data_last_row + 1}"] = f"=SUM({total_col_letter}2:{total_col_letter}{data_last_row})"
            worksheet[f"{label_col_letter}{data_last_row + 1}"].font = Font(bold=True)
            worksheet[f"{total_col_letter}{data_last_row + 1}"].font = Font(bold=True)
            # ===== Format column =====
            for cell in worksheet[total_col_letter]:
                cell.number_format = "0.00"

            # ===== Auto width =====
            for i, col in enumerate(df_transformed.columns):
                max_length = len(str(col))
                for value in df_transformed[col]:
                    max_length = max(max_length, len(str(value)))

                worksheet.column_dimensions[get_column_letter(i + 1)].width = min(max_length + 2, 50)

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

    output.seek(0)
    return output